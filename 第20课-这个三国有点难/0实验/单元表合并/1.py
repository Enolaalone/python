import openpyxl
wb=openpyxl.Workbook()
sheet=wb.create_sheet('表.xlsx')
wb.save('1猜.xlsx')



wb=openpyxl.load_workbook('1.xlsx')
sheet1=wb['表.xlsx']
sheet1.merge_cells(start_row=1,end_row=1,start_column=1,end_column=1+2)#和并表格
wb.save('1猜.xlsx')