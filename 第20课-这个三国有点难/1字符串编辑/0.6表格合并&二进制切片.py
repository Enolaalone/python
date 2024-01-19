import openpyxl

wb=openpyxl.Workbook()
sheet=wb.create_sheet('文本4.xlsx')

A='13!@#$%^&国aA'#使用字符串
def jie(x):
    a=bin(ord(x))
    num=(len(a)-2)%8#看和一字节差几位
    #print(num)
    a1,a2=a.split('b')
    #print(a1,a2)
    a1+='b'#还原0b前缀
    for i in range(8-num):#补充成整字节
        a1 += '0'
    a=a1+a2
    print(a)
    return a

a='国'

def jie_1(a):
    return int(((len(jie(a))-2)/8))

print(jie_1(a))

#必须是创建
sheet.merge_cells(start_row=1,start_column=1,end_row=1,end_column=0+jie_1(a))#和并表格

wb.save('文本4.xlsx')