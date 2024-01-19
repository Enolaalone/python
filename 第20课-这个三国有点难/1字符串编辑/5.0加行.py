import openpyxl as o
#表格读入
wb=o.load_workbook('文本3.xlsx')
sheet=wb['文本3.xlsx']

print(sheet.cell(row=1,column=1).value)
sheet.insert_cols(sheet.max_column)#增加最大列
for i in range(5):
    sheet.cell(row=1,column=1+i,value='123')

wb.save('文本3.xlsx')