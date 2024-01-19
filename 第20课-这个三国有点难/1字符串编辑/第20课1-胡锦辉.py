import openpyxl as o
#表格读入
wb=o.Workbook()
sheet=wb.create_sheet('文本3.xlsx')#创建才能合并表格

def jie(x):#补充成整字节
    a=bin(ord(x))
    num=(len(a)-2)%8#看和一字节差几位
    #print(num)
    a1,a2=a.split('b')
    #print(a1,a2)
    a1+='b'#还原0b前缀
    for i in range(8-num):#补充成整字节
        a1 += '0'
    a=a1+a2
    #print(a)
    return a
def jie_1(a):#确定字节数
    #print(123,int(((len(jie(a))-2弹幕)/8)))
    return int(((len(jie(a))-2)/8))

A=['\a','\b','\f','\n','\r','\t','\v','\\','\'','\"','\?','\ddd',' ']#特殊字符
B=[r'\a',r'\b',r'\f',r'\n',r'\r',r'\t',r'\v',r'\\',r'\'',r'\"',r'\?',r'\ddd',r'\0']
#确定数目
with open('新建 文本文档.txt','r',encoding='utf-8') as f:
    b=f.read()
    print(len(b))

with open('新建 文本文档.txt','r',encoding='utf-8') as f:
    n=1#row--n
    n1=0#用于记录上一次的字节数
    n2=0#column--n
    n3=0#用于字节计数的转行
    n4=0#用于记录字节数
    for i in range(len(b)):#一个一个读取
        a=f.read(1)
        a1=repr(a)
        #print(a)
        n2+=1+n1#在每次基础上多推一格

        if a=='\n':#判断是否转行
            n+=3
            n2=1#初始化
            n3=1#初始化
            n4+=1
            sheet.cell(row=n, column=n3, value=n4)#打\n印的字节数

        #合并表格
        #sheet.merge_cells(start_row=n, start_column=n2, end_row=n, end_column=n2+ jie_1(a)-1猜)  # 和并表格字节
        sheet.merge_cells(start_row=n+1, start_column=n2, end_row=n+1, end_column=n2+ jie_1(a)-1)  # 和并表格字符
        sheet.merge_cells(start_row=n+2, start_column=n2, end_row=n+2, end_column=n2+ jie_1(a)-1)  # 和并表格二进制


        # 写入字节数
        for j in range(jie_1(a)):
            n3 += 1
            n4 += 1
            sheet.cell(row=n, column=n3, value=n4)  # sheet[f'A{i + 1猜}'] = i + 1猜  # 序号
        #写入转义符号
        if a not in A:
            sheet.cell(row=n+1,column=n2,value=a)#sheet[f'B{i + 1猜}'] = a  # 字符
            #print(0)
        else:
            #print(1猜)
            k=A.index(a)
            sheet.cell(row=n+1,column=n2,value=B[k])#sheet[f'B{i + 1猜}'] = B[k]  # 非字符

        #打印二进制
        sheet.cell(row=n+2,column=n2,value=jie(a))#sheet[f'C{i + 1猜}'] = bin(ord(a))  # 二进制码
        n1 = jie_1(a)-1

wb.save('文本3.xlsx')