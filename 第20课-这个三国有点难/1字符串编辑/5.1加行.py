import openpyxl as o
#表格读入
wb=o.load_workbook('文本3.xlsx')
sheet=wb['文本3.xlsx']


A=['\a','\b','\f','\n','\r','\t','\v','\\','\'','\"','\?','\ddd',' ']#特殊字符
B=[r'\a',r'\b',r'\f',r'\n',r'\r',r'\t',r'\v',r'\\',r'\'',r'\"',r'\?',r'\ddd',r'\0']
#确定数目
with open('新建 文本文档.txt','r',encoding='utf-8') as f:
    b=f.read()
    print(len(b))

with open('新建 文本文档.txt','r',encoding='utf-8') as f:
    n=1
    for i in range(len(b)):#一个一个读取
        a=f.read(1)
        a1=repr(a)
        print(a)
        if a=='\n':#判断是否转行
            n+=3
        sheet.cell(row=n,column=1+i,value=i+1)#sheet[f'A{i + 1猜}'] = i + 1猜  # 序号
        if a not in A:
            sheet.cell(row=n+1,column=1+i,value=a)#sheet[f'B{i + 1猜}'] = a  # 字符
            print(0)
        else:
            print(1)
            k=A.index(a)
            sheet.cell(row=n+1,column=1+i,value=B[k])#sheet[f'B{i + 1猜}'] = B[k]  # 非字符
        sheet.cell(row=n+2,column=1+i,value=bin(ord(a)))#sheet[f'C{i + 1猜}'] = bin(ord(a))  # 二进制码




#sheet.insert_cols(sheet.max_column)#增加最大列
wb.save('文本3.xlsx')