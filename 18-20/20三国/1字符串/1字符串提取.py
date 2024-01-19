import openpyxl


def judge(n):
    return len(n)-2


with open('新建 文本文档.txt','r',encoding='utf-8') as f:
    text = f.readlines()
wb=openpyxl.Workbook()
sheet=wb.create_sheet('字符串.xlsx')
n2=-1
for sentence in text:
    n2+=2
    n1=0
    n3=0
    for word in sentence:
        n1+=1+n3
        char=bin(ord(word))
        n3=judge(char)//8
        if word=='\n':
            word=r'\n'
        sheet.merge_cells(start_row=n2,end_row=n2,start_column=n1,end_column=n1+n3)
        sheet.cell(row=n2,column=n1,value=word)
        sheet.merge_cells(start_row=n2+1,end_row=n2+1,start_column=n1,end_column=n1+n3)
        sheet.cell(row=n2+1,column=n1,value=char)
wb.save('字符串.xlsx')