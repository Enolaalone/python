import Data2
import openpyxl


def deal_time(n):
    r_time=n
    t=r_time[-8:].split(':')
    r_time=float(t[0])+float(t[1])/60+float(t[2])/3600
    return r_time

f=open('实验室名单.txt','r',encoding='utf-8')
names=f.readlines()
for name in names:
    name=name.strip('\n')
    Data2.names[name]=0
f.close()
print(Data2.names)

wb = openpyxl.load_workbook('10课.xlsx')
sheet=wb['成员参会概况']
r_time=deal_time(sheet["B4"].value)
for i in range(10,sheet.max_row):
    if deal_time(sheet[f'B{i}'].value)>r_time:
        Data2.nickname.append(sheet[f'A{i}'].value)
print(Data2.nickname)

for i in range(10,sheet.max_row):
    Data2.whole_nickname.append(sheet[f'A{i}'].value)
for i in Data2.whole_nickname:
    for key in Data2.names:
        if key in i:
            Data2.names[key]=2
print(Data2.names)

for i in Data2.nickname:
    for key in Data2.names:
        if key in i:
            Data2.names[key]=1

print(Data2.names)

list=list(Data2.names)
for i in list:
    if Data2.names[i]==2:
        del Data2.names[i]
print(Data2.names)