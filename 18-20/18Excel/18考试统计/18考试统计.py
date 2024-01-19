import Data3
import openpyxl


def grade(n):
    grade=int(n.strip('秒'))
    if grade-180<=0:
        return 0
    else:
        return (grade-180)/2


wb1 = openpyxl.load_workbook('班级名单.xlsx')
wb2 = openpyxl.load_workbook('导论考试01--进制转换2.xlsx')
sheet1 = wb1['导论655686']
sheet2 = wb2['Sheet1']

for row in sheet1.iter_rows(min_row=2,max_row=sheet1.max_row,max_col=3,min_col=3):
    Data3.names[row[0].value]=0
print(Data3.names)

for row in sheet2.iter_rows(min_row=2,max_row=sheet2.max_row,min_col=2,max_col=2):
    for name in Data3.names:
        if name in row[0].value:
            row[0].value=name

for row in sheet2.iter_rows(min_row=2,max_row=sheet2.max_row,min_col=2,max_col=4):
    grades_0=row[1].value
    grades_1=grades_0-grade(row[2].value)
    if grades_1>Data3.names[row[0].value]:
        Data3.names[row[0].value]=grades_1
print(Data3.names)
