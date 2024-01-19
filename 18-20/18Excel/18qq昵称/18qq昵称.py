import Data1
import openpyxl

wb1=openpyxl.load_workbook('班级名单.xlsx')
wb2=openpyxl.load_workbook('QQ群昵称总表.xlsx')
sheet1=wb1['导论682006']
sheet2=wb2['导论862006']
for i in sheet1.iter_rows(min_row=2,max_row=sheet1.max_row):
    Data1.name[i[2].value]=str(i[1].value)
print(Data1.name)
for i in range(2,sheet1.max_row+1):
    if sheet2[f"D{i}"].value!=None:
        Data1.nickname.append(sheet2[f'D{i}'].value)
print(Data1.nickname)

for key,value in Data1.name.items():
    perfect=value+key
    n=0
    for i in Data1.nickname:
        n+=1
        if i==perfect:
            Data1.name[key]=100
            break
        if key and value in i:
            Data1.name[key]=90
            break
        if key in i:
            Data1.name[key]=80
            break
        if n==len(Data1.nickname):
            Data1.name[key]=0

n=1
for i in Data1.name:
    n+=1
    sheet1[f'F{n}']=Data1.name[i]
wb1.save('班级名单.xlsx')