import Data2
import openpyxl


# def read():
#     wb = openpyxl.load_workbook('商店.xlsx')
#     sheet = wb['商品.xlsx']
#     Data2.thing = {}
#     for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
#         Data2.thing[row[0].value] = row[1].value
#         Data2.thing[row[0].value] += f':{row[2].value}'
#         # print(Data2.thing[row[0].value])
#         for i in range(3):
#             row[i].value = None
#     wb.save('商店.xlsx')
#     print(Data2.thing)
#
#
# def record():
#     wb = openpyxl.load_workbook('商店.xlsx')
#     sheet = wb['商品.xlsx']
#     n = 1
#     print(Data2.thing)
#     for key, value in Data2.thing.items():
#         n += 1
#         print(value)
#         a, b = value.split(':')
#         sheet[f'A{n}'] = key
#         sheet[f'B{n}'] = a
#         sheet[f'C{n}'] = b
#     wb.save('商店.xlsx')
#
#
# def change(n, num):
#     a, b = n.split(':')
#     if num == 1:
#         a = input('输入商品名称')
#         a += f':{b}'
#         return a
#     if num == 2:
#         b = input('输入修改价格')
#         a += f':{b}元/个'
#         return a
#
#
# def change_plus(n):
#     while True:
#         num = int(input('输入商品编码'))
#         if num not in Data2.thing:
#             print("商品不存在")
#         else:
#             break
#     Data2.thing[num] = change(Data2.thing[num], n)
#     print(Data2.thing)

# def price(n):
#     a, b = n.split(':')
#     b = b[0:-3]
#     # print(f'价格{b}')
#     return int(b)
#
# def bill():
#     wb=openpyxl.load_workbook('商店.xlsx')
#     sheet=wb.create_sheet(f'账单.xlsx')
#     sheet["A1"]='商品'
#     sheet["B1"]='数量'
#     sheet["C1"]='价格'
#
#     money = 0
#     i=1
#     for key, value in Data2.subject.items():
#         i+=1
#         Data2.subject[key] = price(key) * value
#         sheet[f'A{i}']=f'{key}'
#         sheet[f'B{i}']=f'{value}'
#         sheet[f'C{i}']=f'{Data2.subject[key]}'
#         money += Data2.subject[key]
#     sheet[f'C{i+1}']=f'共计{money}元'
#     wb.save('商店.xlsx')
#
# def hot_read():
#     wb = openpyxl.load_workbook('商店.xlsx')
#     sheet = wb['热门.xlsx']
#     Data2.hot = {}
#     for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
#         Data2.hot[row[0].value] = row[1].value
#         # print(Data2.thing[row[0].value])
#         for i in range(2):
#             row[i].value = None
#     wb.save('商店.xlsx')
#
# def hot_record():
#     Data2.re=sorted(Data2.hot, key=Data2.hot.get, reverse=True)
#     print(Data2.re)
#     wb = openpyxl.load_workbook('商店.xlsx')
#     sheet = wb['热门.xlsx']
#     n = 1
#     for key in Data2.re:
#         n += 1
#         sheet[f'A{n}'] = key
#         sheet[f'B{n}'] = Data2.hot[key]
#     wb.save('商店.xlsx')
#
# ----------------------------------------------------------------------
# read()
# print(Data2.thing)
# while True:
#     while True:
#         num=int(input('输入商品编码'))
#         if num in Data2.thing:
#             print('商品已经存在')
#         else:
#             break
#     name=input('输入商品名称')
#     price0=input('商品价格')
#     Data2.thing[num]=f'{name}:{price0}元/个'
#     print(Data2.thing)
#     a=input('Enter继续输入  w退出')
#     if a=='w':
#         record()
#         break


# ------------------------------------------------------------------------
# read()
# print(Data2.thing)
# num=int(input('输入商品编码'))
# if num in Data2.thing:
#     print(f'{Data2.thing[num]}')
# else:
#     print('商品不存在')


# -------------------------------------------------------------------------
# read()
# print(Data2.thing)
# while True:
#     while True:
#         num = int(input('商品编码'))
#         if num not in Data2.thing:
#
#             print('商品不存在')
#         else:
#             break
#     del Data2.thing[num]
#     print(Data2.thing)
#     a = input('Enter继续删除  w退出')
#     if a == 'w':
#         record()
#         break
# ----------------------------------------------------------------------------
# read()
# while True:
#     b=input('1修改商品名称    2修改价格')
#     print(Data2.thing)
#     if b=='1':
#         change_plus(1)
#         a=input('Enter继续输入  w退出')
#         if a=='w':
#             record()
#             break
#     if b=='2':
#         change_plus(2)
#         a=input('Enter继续输入  w退出')
#         if a=='w':
#             record()
#             break
# --------------------------------------------------------------------------

# read()
# hot_read()
# while True:
#     while True:
#         num=int(input('商品编码'))
#         if num not in Data2.thing:
#             print('无此商品')
#         else:
#             break
#
#     print(f'{Data2.thing[num]}x1')
#
#     if Data2.thing[num] not in Data2.hot:
#         Data2.hot[Data2.thing[num]] = 1
#     else:
#         Data2.hot[Data2.thing[num]] += 1
#
#     if Data2.thing[num] not in Data2.subject:
#         Data2.subject[Data2.thing[num]]=1
#     else:
#         Data2.subject[Data2.thing[num]]+=1
#     a=input('Enter继续    w结账')
#     if a=='w':
#         bill()
#         record()
#         hot_record()
#         break