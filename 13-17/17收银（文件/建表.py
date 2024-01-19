import openpyxl
#utf-8格式建表
wb=openpyxl.Workbook()
sheet0 = wb.create_sheet('商品.xlsx')
sheet1 = wb.create_sheet('员工信息.xlsx')
sheet2 = wb.create_sheet('账单.xlsx')
sheet3=wb.create_sheet('热门.xlsx')
wb.save('商店.xlsx')