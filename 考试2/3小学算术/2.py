import random
import openpyxl as o
wb=o.Workbook()
sheet1=wb.create_sheet('sheet.xlsx')
s='y'
n1=0
while s=='y':
    n1+=1
    name=input('姓名')
    sheet1['A1']='姓名'
    sheet1[f'A{n1+1}']=f'{name}'
    sheet=wb.create_sheet(f'{name}.xlsx')
    sheet['A1']='题目'
    sheet['B1']='答案'
    sheet['C1']=f'{name}答案'
    sheet['D1']='单题得分'
    sheet['F1']='加法题数'
    sheet['G1']='减法题数'
    sheet['H1']='乘法题数'
    sheet['I1']='除法题数'
    sheet['J1']='个位数加法'
    A={}
    def jia_g():
        a = random.randint(0, 9)
        b = random.randint(0, 9)
        a1 = str(a)
        b1 = str(b)
        a1 += '+'
        a1 += b1
        a1 += '='
        A[a + b] = a1
    def jia():
        a =random.randint(0, 99)
        b =random.randint(0, 100-a)
        a1=str(a)
        b1=str(b)
        a1 += '+'
        a1 += b1
        a1 += '='
        A[a+b] = a1

    def jian():
        a = random.randint(0, 99)

        b = random.randint(0, a)
        a1=str(a)
        b1=str(b)
        a1 += '-'
        a1 += b1
        a1 += '='
        A[a - b] = a1

    def cheng():
        while True:
            a = random.randint(0, 99)
            b = random.randint(0, 10)
            if a*b<100:
                break
        a1=str(a)
        b1=str(b)
        a1+='X'
        a1 += b1
        a1 += '='
        A[a * b] = a1

    def chu():
        while True:
            a = random.randint(0, 99)
            b = random.randint(1, a)
            if a%b==0:
                break
            else:
                continue
        a1=str(a)
        b1=str(b)
        a1 += '/'
        a1 += b1
        a1 += '='
        A[int(a/b)] = a1
    a1=0
    a2=0
    a3=0
    a4=0
    a5=0
    for j in range(2):
        k=random.randint(1,4)
        if k==1:
            jian()
            a1+=1
        if k==2:
            jian()
            a2 += 1
        if k==3:
            cheng()
            a3 += 1
        if k==4:
            chu()
            a4+=1
    for i in range(2):
        k1=random.randint(1,2)
        if k1==1:
            jia_g()
            a5+=1
        else:
            jia()
        jian()
        cheng()
        chu()
    sheet['F2']=2+a1
    sheet['G2']=2+a2
    sheet['H2']=2+a3
    sheet['I2']=2+a4
    sheet['J2'] =a5
    print(A)
    n=2
    for key,value in A.items():
        sheet[f'A{n}']=value
        sheet[f'B{n}']=key
        n+=1
        print(value)

    B=[]
    n=2
    for i in range(10):
        a=int(input('输入结果'))
        sheet[f'C{n}'] = a
        B.append(a)
        n+=1

    grade=0
    n=2
    for key,value in A.items():
        if B[n-2]==key:
            sheet[f'D{n}']=10
            grade+=10
        else:
            sheet[f'D{n}']=0
        n+=1

    sheet[f'E1']='成绩'
    sheet[f'E2']=grade
    sheet1[f'E1'] = '成绩'
    sheet1[f'E{n1+1}'] = grade
    if grade<60:
        sheet[f'E3'] = '不及格'
    if 60<=grade<70:
        sheet[f'E3'] = '及格'
    if 70<=grade<80:
        sheet[f'E3'] = '一般'
    if 80<=grade<90:
        sheet[f'E3'] = '良好'
    if 90<=grade<=100:
        sheet[f'E3'] = '优秀'
    print(grade)

    wb.save('成绩.xlsx')
    s=input('下一个人是否继续？y/n')